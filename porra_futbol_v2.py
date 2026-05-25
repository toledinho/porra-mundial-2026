import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import os

def analizar_predicciones(datos_brutos):
    """Analiza los datos brutos de predicciones en un formato estructurado."""
    lineas = datos_brutos.strip().split('\n')
    participantes = []
    
    for linea in lineas:
        partes = linea.split()
        if len(partes) < 4:  # Necesitamos al menos nombre y 3 predicciones
            continue
            
        # El nombre puede tener múltiples palabras, así que necesitamos encontrar dónde comienzan las predicciones
        # Las predicciones están en formato X-X
        indice_inicio_pred = 0
        for i, parte in enumerate(partes):
            if '-' in parte and parte.replace('-', '').isdigit():
                indice_inicio_pred = i
                break
        
        if indice_inicio_pred == 0:
            continue
            
        nombre = ' '.join(partes[:indice_inicio_pred])
        predicciones = partes[indice_inicio_pred:indice_inicio_pred+3]
        
        # Añadir una cuarta predicción como None si no se proporciona
        if len(predicciones) < 4:
            predicciones.append(None)
            
        participantes.append({
            'nombre': nombre,
            'pred1': predicciones[0] if len(predicciones) > 0 else None,
            'pred2': predicciones[1] if len(predicciones) > 1 else None,
            'pred3': predicciones[2] if len(predicciones) > 2 else None,
            'pred4': predicciones[3] if len(predicciones) > 3 else None
        })
    
    return participantes

def calcular_puntos(prediccion, resultado_real):
    """
    Calcula los puntos basados en las reglas de puntuación:
    - 12 puntos: Partido acertado con diferencia de goles > 1
    - 10 puntos: Partido acertado con diferencia de goles = 1 o empate
    - 6 puntos: Partido no acertado pero ganador acertado con diferencia de goles
    - 4 puntos: Partido no acertado, diferencia de goles no acertada pero ganador acertado
    - 0 puntos: Ganador no acertado
    """
    if prediccion is None:
        return 0
        
    pred_local, pred_visitante = map(int, prediccion.split('-'))
    real_local, real_visitante = map(int, resultado_real.split('-'))
    
    # Resultado exacto
    if pred_local == real_local and pred_visitante == real_visitante:
        # Comprobar diferencia de goles
        dif_goles = abs(real_local - real_visitante)
        if dif_goles > 1:
            return 12  # Resultado exacto con diferencia de goles > 1
        else:
            return 10  # Resultado exacto con diferencia de goles = 1 o empate
    
    # Calcular resultados previstos y reales
    if pred_local > pred_visitante:
        resultado_previsto = 'local'
    elif pred_local < pred_visitante:
        resultado_previsto = 'visitante'
    else:
        resultado_previsto = 'empate'
        
    if real_local > real_visitante:
        resultado_real_ganador = 'local'
    elif real_local < real_visitante:
        resultado_real_ganador = 'visitante'
    else:
        resultado_real_ganador = 'empate'
    
    # Comprobar si el ganador es correcto
    if resultado_previsto == resultado_real_ganador:
        # Comprobar si la diferencia de goles es correcta
        dif_prevista = pred_local - pred_visitante
        dif_real = real_local - real_visitante
        
        if dif_prevista == dif_real:
            return 6  # Resultado incorrecto pero ganador y diferencia de goles correctos
        else:
            return 4  # Resultado incorrecto, diferencia de goles incorrecta pero ganador correcto
    
    return 0  # Ganador incorrecto

def obtener_numero_jornada():
    """Obtiene el número de la jornada actual."""
    try:
        with open('numero_jornada.txt', 'r', encoding='utf-8') as f:
            numero = int(f.read().strip())
            return numero
    except FileNotFoundError:
        return 1

def guardar_numero_jornada(numero):
    """Guarda el número de la jornada actual."""
    with open('numero_jornada.txt', 'w', encoding='utf-8') as f:
        f.write(str(numero))

def crear_excel(participantes, resultados_reales, nombres_partidos, archivo_salida='porra_futbol.xlsx', exportar_imagen=True):
    """Crea un archivo Excel con los resultados y opcionalmente exporta como imagen."""
    
    # Calcular puntos para cada participante y partido
    for participante in participantes:
        participante['puntos1'] = calcular_puntos(participante['pred1'], resultados_reales[0])
        participante['puntos2'] = calcular_puntos(participante['pred2'], resultados_reales[1])
        participante['puntos3'] = calcular_puntos(participante['pred3'], resultados_reales[2])
        
        # Manejar el 4º partido (puntos dobles si se proporciona)
        if len(resultados_reales) > 3 and participante['pred4'] is not None:
            participante['puntos4'] = calcular_puntos(participante['pred4'], resultados_reales[3]) * 2
        else:
            participante['puntos4'] = 0
            
        # Calcular totales
        participante['total'] = participante['puntos1'] + participante['puntos2'] + participante['puntos3'] + participante['puntos4']
    
    # Crear DataFrame
    df = pd.DataFrame(participantes)
    
    # Ordenar por puntos totales (descendente)
    df = df.sort_values('total', ascending=False)
    
    # Obtener el número de jornada
    numero_jornada = obtener_numero_jornada()
    
    # Verificar si el archivo Excel ya existe
    if os.path.exists(archivo_salida):
        # Cargar el archivo existente
        wb = load_workbook(archivo_salida)
    else:
        # Crear un nuevo libro de Excel
        wb = Workbook()
        # Eliminar la hoja por defecto
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # Crear una nueva hoja para la jornada actual
    ws = wb.create_sheet(f"Jornada {numero_jornada}")
    
    # Añadir encabezado
    encabezado = ["Participante", 
              f"{nombres_partidos[0]} ({resultados_reales[0]})", "Predicción 1", "Puntos 1",
              f"{nombres_partidos[1]} ({resultados_reales[1]})", "Predicción 2", "Puntos 2",
              f"{nombres_partidos[2]} ({resultados_reales[2]})", "Predicción 3", "Puntos 3"]
              
    # Añadir 4º partido si está disponible
    if len(resultados_reales) > 3 and len(nombres_partidos) > 3:
        encabezado.extend([f"{nombres_partidos[3]} ({resultados_reales[3]})", "Predicción 4", "Puntos 4"])
        
    encabezado.extend(["Total Jornada", "Total Mes", "Total Temporada"])
    
    ws.append(encabezado)
    
    # Añadir datos
    for _, fila in df.iterrows():
        fila_datos = [fila['nombre'],
                   resultados_reales[0], fila['pred1'], fila['puntos1'],
                   resultados_reales[1], fila['pred2'], fila['puntos2'],
                   resultados_reales[2], fila['pred3'], fila['puntos3']]
                   
        # Añadir 4º partido si está disponible
        if len(resultados_reales) > 3 and len(nombres_partidos) > 3:
            fila_datos.extend([resultados_reales[3], fila['pred4'], fila['puntos4']])
            
        # Añadir totales (jornada, mes y temporada)
        total_jornada = fila['total']
        total_mes = fila.get('total_mes', 0) + total_jornada
        total_temporada = fila.get('total_temporada', 0) + total_jornada
        
        fila_datos.extend([total_jornada, total_mes, total_temporada])
        
        ws.append(fila_datos)
    
    # Dar formato al Excel
    # Definir estilos
    encabezado_estilo = Font(bold=True)
    borde = Border(left=Side(style='thin'), 
                 right=Side(style='thin'), 
                 top=Side(style='thin'), 
                 bottom=Side(style='thin'))
    alineacion_centro = Alignment(horizontal='center')
    
    # Aplicar estilos
    for fila in ws.iter_rows(min_row=1, max_row=1):
        for celda in fila:
            celda.font = encabezado_estilo
            celda.border = borde
            celda.alignment = alineacion_centro
    
    # Aplicar bordes y alineación a todas las celdas con datos
    for fila in ws.iter_rows(min_row=2, max_row=len(participantes) + 1):
        for celda in fila:
            celda.border = borde
            celda.alignment = alineacion_centro
    
    # Ajustar ancho de columnas
    for i, columna in enumerate(ws.columns, 1):
        max_length = 0
        columna_letra = get_column_letter(i)
        
        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))
        
        adjusted_width = max_length + 2
        ws.column_dimensions[columna_letra].width = adjusted_width
    
    # Guardar el archivo
    wb.save(archivo_salida)
    print(f"Archivo Excel guardado como {archivo_salida}")
    
    # Incrementar y guardar el número de jornada para la próxima vez
    guardar_numero_jornada(numero_jornada + 1)
    
    # Exportar como imagen (PDF)
    if exportar_imagen:
        # Crear un DataFrame para la visualización
        df_visual = pd.DataFrame()
        df_visual['Participante'] = df['nombre']
        
        # Añadir columnas para cada partido
        for i in range(3):
            df_visual[f"{nombres_partidos[i]} ({resultados_reales[i]})"] = df[f'pred{i+1}']
            df_visual[f"Puntos {i+1}"] = df[f'puntos{i+1}']
            
        # Añadir 4º partido si está disponible
        if len(resultados_reales) > 3 and len(nombres_partidos) > 3:
            df_visual[f"{nombres_partidos[3]} ({resultados_reales[3]})"] = df['pred4']
            df_visual[f"Puntos 4"] = df['puntos4']
            
        # Añadir totales
        df_visual["Total Jornada"] = df['total']
        df_visual["Total Mes"] = df['total']
        df_visual["Total Temporada"] = df['total']
        
        # Crear PDF
        pdf_filename = f'jornada{numero_jornada}.pdf'
        with PdfPages(pdf_filename) as pdf:
            # Configurar figura
            fig, ax = plt.subplots(figsize=(15, len(participantes) * 0.4 + 2))
            ax.axis('tight')
            ax.axis('off')
            
            # Crear tabla
            tabla = ax.table(
                cellText=df_visual.values,
                colLabels=df_visual.columns,
                loc='center',
                cellLoc='center'
            )
            
            # Ajustar estilo de la tabla
            tabla.auto_set_font_size(False)
            tabla.set_fontsize(8)
            tabla.scale(1.5, 1.8)
            
            # Ajustar el ancho de las columnas
            # Hacer la primera columna (nombres) más ancha
            for i in range(len(df_visual.columns)):
                if i == 0:  # Columna de nombres
                    tabla.auto_set_column_width(i)
                    # Hacer la columna de nombres un 50% más ancha
                    col_width = tabla._approx_text_height() * 1.5
                    for cell in tabla._cells:
                        if cell[1] == 0:  # Primera columna
                            tabla._cells[cell].set_width(col_width)
            
            # Ajustar el título
            plt.title(f'Clasificación Porra de Fútbol - Jornada {numero_jornada}', fontsize=16, pad=20)
            
            # Guardar como PDF
            pdf.savefig(fig, bbox_inches='tight')
            plt.close()
            
        print(f"Imagen guardada como {pdf_filename}")

def solicitar_nombres_partidos(num_partidos):
    """Solicita al usuario los nombres de los partidos o carga los guardados previamente."""
    # Comprobar si existe un archivo con los nombres de los partidos
    archivo_nombres = 'nombres_partidos.txt'
    try:
        with open(archivo_nombres, 'r', encoding='utf-8') as f:
            nombres_partidos = [line.strip() for line in f.readlines()]
            if len(nombres_partidos) >= num_partidos:
                print(f"Usando nombres de partidos guardados previamente:")
                for i, nombre in enumerate(nombres_partidos[:num_partidos]):
                    print(f"Partido {i+1}: {nombre}")
                return nombres_partidos[:num_partidos]
    except FileNotFoundError:
        pass
    
    # Si no existe el archivo o no tiene suficientes nombres, solicitar al usuario
    nombres_partidos = []
    for i in range(num_partidos):
        nombre = input(f"Introduce el nombre del partido {i+1} (Ej: Equipo Local - Equipo Visitante): ")
        nombres_partidos.append(nombre)
    
    # Guardar los nombres para futuras ejecuciones
    with open(archivo_nombres, 'w', encoding='utf-8') as f:
        for nombre in nombres_partidos:
            f.write(f"{nombre}\n")
    
    return nombres_partidos

def cargar_puntuaciones_previas():
    """Carga las puntuaciones acumuladas de jornadas anteriores."""
    try:
        with open('puntuaciones_acumuladas.txt', 'r', encoding='utf-8') as f:
            puntuaciones = {}
            for linea in f:
                partes = linea.strip().split(',')
                if len(partes) >= 3:
                    nombre = partes[0]
                    total_mes = int(partes[1])
                    total_temporada = int(partes[2])
                    puntuaciones[nombre] = {
                        'total_mes': total_mes,
                        'total_temporada': total_temporada
                    }
            return puntuaciones
    except FileNotFoundError:
        return {}

def guardar_puntuaciones(participantes, es_primer_partido_mes):
    """Guarda las puntuaciones acumuladas para futuras jornadas."""
    with open('puntuaciones_acumuladas.txt', 'w', encoding='utf-8') as f:
        for participante in participantes:
            nombre = participante['nombre']
            total_jornada = participante['total']
            
            # Si es el primer partido del mes, reiniciar el total mensual
            if es_primer_partido_mes:
                total_mes = total_jornada
            else:
                total_mes = participante.get('total_mes', 0) + total_jornada
            
            # Acumular el total de la temporada
            total_temporada = participante.get('total_temporada', 0) + total_jornada
            
            f.write(f"{nombre},{total_mes},{total_temporada}\n")

def solicitar_resultados():
    """Solicita al usuario los resultados reales de los partidos."""
    resultados = []
    for i in range(3):
        while True:
            resultado = input(f"Introduce el resultado del partido {i+1} (formato: X-X): ")
            if '-' in resultado and resultado.replace('-', '').isdigit():
                resultados.append(resultado)
                break
            else:
                print("Formato incorrecto. Por favor, introduce el resultado en formato X-X (ejemplo: 2-1)")
    
    # Preguntar si hay partido estrella
    while True:
        respuesta = input("¿Hay partido estrella? (s/n): ").lower()
        if respuesta in ['s', 'si', 'sí', 'y', 'yes']:
            while True:
                resultado = input("Introduce el resultado del partido estrella (formato: X-X): ")
                if '-' in resultado and resultado.replace('-', '').isdigit():
                    resultados.append(resultado)
                    break
                else:
                    print("Formato incorrecto. Por favor, introduce el resultado en formato X-X (ejemplo: 2-1)")
            break
        elif respuesta in ['n', 'no']:
            break
        else:
            print("Por favor, responde 's' o 'n'")
    
    return resultados

def solicitar_pronosticos():
    """Solicita al usuario que introduzca los pronósticos de los participantes."""
    print("Por favor, introduce los pronósticos de los participantes.")
    print("Formato: NOMBRE X-X X-X X-X (un participante por línea)")
    print("Ejemplo: JUAN 1-0 2-1 0-0")
    print("Escribe 'FIN' en una línea para terminar.")
    
    datos_brutos = ""
    while True:
        linea = input("> ")
        if linea.upper() == 'FIN':
            break
        datos_brutos += linea + "\n"
    
    return datos_brutos

def main():
    # Solicitar pronósticos de los participantes
    datos_brutos = solicitar_pronosticos()
    
    # Solicitar resultados reales
    print("Por favor, introduce los resultados reales de los partidos:")
    resultados_reales = solicitar_resultados()
    
    # Preguntar si es el primer partido del mes
    while True:
        respuesta = input("¿Es el primer partido del mes? (s/n): ").lower()
        if respuesta in ['s', 'si', 'sí', 'y', 'yes']:
            es_primer_partido_mes = True
            break
        elif respuesta in ['n', 'no']:
            es_primer_partido_mes = False
            break
        else:
            print("Por favor, responde 's' o 'n'")
    
    # Cargar puntuaciones previas
    puntuaciones_previas = cargar_puntuaciones_previas()
    
    # Determinar el número de partidos
    num_partidos = len(resultados_reales)
    
    # Solicitar nombres de partidos
    print("Por favor, introduce los nombres de los partidos:")
    nombres_partidos = solicitar_nombres_partidos(num_partidos)
    
    # Analizar predicciones
    participantes = analizar_predicciones(datos_brutos)
    
    # Añadir puntuaciones previas a los participantes
    for participante in participantes:
        nombre = participante['nombre']
        if nombre in puntuaciones_previas:
            participante['total_mes'] = puntuaciones_previas[nombre]['total_mes']
            participante['total_temporada'] = puntuaciones_previas[nombre]['total_temporada']
        else:
            participante['total_mes'] = 0
            participante['total_temporada'] = 0
    
    # Crear Excel y exportar como imagen
    crear_excel(participantes, resultados_reales, nombres_partidos, exportar_imagen=True)
    
    # Guardar puntuaciones acumuladas
    guardar_puntuaciones(participantes, es_primer_partido_mes)

if __name__ == "__main__":
    main()